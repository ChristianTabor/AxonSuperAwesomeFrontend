from openpyxl import load_workbook, Workbook
import datetime
from sql import *


def test():
    # filename = "AxonOutPut-" + date.today().strftime("%b-%d-%Y") + ".xlsx"

    wb = Workbook()
    ws = wb.active
    ws['A1'] = 12
    ws.insert_cols(2)
    ws.insert_rows(1)
    ws['A1'] = 14
    # wb.save(filename)


def excelGetAttendees(sqlVars: [], filename):
    vals = getAllAttendees(sqlVars)
    workbook = load_workbook(filename=filename)
    sheet = workbook.create_sheet("Attendees")
    for i in vals:
        res = eval(str(i))
        sheet.insert_rows(1)
        sheet['A1'] = res[0]
        sheet['B1'] = res[1]
        sheet['C1'] = res[2]
        sheet['D1'] = res[3]
        sheet['E1'] = res[4]
        sheet['F1'] = res[5]
        sheet['G1'] = res[6]
    sheet.insert_rows(1)
    sheet['A1'] = "ID"
    sheet['B1'] = "Name"
    sheet['C1'] = "Email"
    sheet['D1'] = "Phone"
    sheet['E1'] = "Company"
    sheet['F1'] = "Created"
    sheet['G1'] = "Updated"
    workbook.save(filename)


def excelGetBooths(sqlVars: [], filename):
    vals = getAllBooths(sqlVars)
    workbook = load_workbook(filename=filename)
    sheet = workbook.create_sheet("Booths")
    for i in vals:
        res = eval(str(i))
        sheet.insert_rows(1)
        sheet['A1'] = res[0]
        sheet['B1'] = res[1]
        sheet['C1'] = res[2]
        sheet['D1'] = res[3]
    sheet.insert_rows(1)
    sheet['A1'] = "ID"
    sheet['B1'] = "Name"
    sheet['C1'] = "Created"
    sheet['D1'] = "Updated"
    workbook.save(filename)


def excelGetEvents(sqlVars: [], filename):
    vals=getAllEvents(sqlVars)
    workbook = load_workbook(filename=filename)
    sheet = workbook.create_sheet("Events")
    for i in vals:
        res = eval(str(i))
        sheet.insert_rows(1)
        sheet['A1'] = res[0]
        sheet['B1'] = res[1]
        sheet['C1'] = res[2]
        sheet['D1'] = res[3]
        sheet['E1'] = res[4]
        sheet['F1'] = res[5]
    sheet.insert_rows(1)
    sheet['A1'] = "ID"
    sheet['B1'] = "Name"
    sheet['C1'] = "Location"
    sheet['D1'] = "ZIP Code"
    sheet['E1'] = "Created"
    sheet['F1'] = "Updated"
    workbook.save(filename)


def excelGetInteractions(sqlVars: [], filename):
    vals = getAllInteractions(sqlVars)
    print(vals)
    workbook = load_workbook(filename=filename)
    sheet = workbook.create_sheet("Scans")
    for i in vals:
        res = eval(str(i))
        sheet.insert_rows(1)
        sheet['A1'] = res[0]
        sheet['B1'] = res[1]
        sheet['C1'] = res[2]
        sheet['D1'] = res[3]
        sheet['E1'] = res[4]
        sheet['F1'] = res[5]
        sheet['G1'] = res[6]
        sheet['H1'] = res[7]
    sheet.insert_rows(1)
    sheet['A1'] = "Name"
    sheet['B1'] = "Email"
    sheet['C1'] = "Phone"
    sheet['D1'] = "Company"
    sheet['E1'] = "Booth Name"
    sheet['F1'] = "Booth ID"
    sheet['G1'] = "Event"
    sheet['H1'] = "Event ID"
    workbook.save(filename)
